import csv
import io
from collections.abc import Iterable

from openpyxl import load_workbook

from asta_la_vista.domain.commands import PlayerRow
from asta_la_vista.exceptions import ValidationError

REQUIRED_COLUMNS = ("Id", "R", "Nome", "Squadra")
QUOTATION_COLUMN = "Qt.A"
MANTRA_ROLE_COLUMN = "RM"


def parse_player_file(content: bytes, filename: str) -> tuple[PlayerRow, ...]:
    extension = filename.rsplit(".", 1)[-1].lower()
    if extension == "xlsx":
        return _parse_xlsx(content)
    if extension == "csv":
        return _parse_csv(content)
    raise ValidationError("Only .xlsx and .csv files are supported")


def _parse_xlsx(content: bytes) -> tuple[PlayerRow, ...]:
    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise ValidationError("The Excel file is not valid") from exc
    if "Tutti" not in workbook.sheetnames:
        raise ValidationError("The Excel file does not contain the 'Tutti' sheet")
    rows = workbook["Tutti"].iter_rows(values_only=True)
    headers, data = _find_headers(rows)
    return _rows_to_players(headers, data)


def _parse_csv(content: bytes) -> tuple[PlayerRow, ...]:
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise ValidationError("The CSV file must use UTF-8 encoding") from exc
    try:
        dialect = csv.Sniffer().sniff(text[:4096], delimiters=",;")
    except csv.Error:
        dialect = csv.excel
    rows = csv.reader(io.StringIO(text), dialect)
    headers, data = _find_headers(rows)
    return _rows_to_players(headers, data)


def _find_headers(rows: Iterable[Iterable[object]]) -> tuple[list[str], Iterable[Iterable[object]]]:
    iterator = iter(rows)
    for row_number, row in enumerate(iterator, start=1):
        values = [str(value).strip() if value is not None else "" for value in row]
        if all(column in values for column in REQUIRED_COLUMNS):
            return values, iterator
        if row_number == 10:
            break
    raise ValidationError("Required columns Id, R, Nome and Squadra were not found")


def _rows_to_players(headers: list[str], rows: Iterable[Iterable[object]]) -> tuple[PlayerRow, ...]:
    positions = {column: headers.index(column) for column in REQUIRED_COLUMNS}
    quotation_position = headers.index(QUOTATION_COLUMN) if QUOTATION_COLUMN in headers else None
    mantra_role_position = (
        headers.index(MANTRA_ROLE_COLUMN) if MANTRA_ROLE_COLUMN in headers else None
    )
    players: list[PlayerRow] = []
    for row in rows:
        values = list(row)
        if not any(value is not None and str(value).strip() for value in values):
            continue
        try:
            external_id = str(values[positions["Id"]]).strip()
            role = str(values[positions["R"]]).strip()
            name = str(values[positions["Nome"]]).strip()
            team = str(values[positions["Squadra"]]).strip()
        except IndexError as exc:
            raise ValidationError("A player row is incomplete") from exc
        if not all((external_id, role, name, team)) or external_id == "None":
            raise ValidationError("A player row contains empty required values")
        quotation = (
            _parse_quotation(values[quotation_position])
            if quotation_position is not None and quotation_position < len(values)
            else None
        )
        mantra_roles = (
            _parse_mantra_roles(values[mantra_role_position])
            if mantra_role_position is not None and mantra_role_position < len(values)
            else (() if mantra_role_position is not None else None)
        )
        players.append(PlayerRow(external_id, name, team, role, quotation, mantra_roles))
    if not players:
        raise ValidationError("The player list is empty")
    return tuple(players)


def _parse_mantra_roles(value: object) -> tuple[str, ...]:
    if value is None or not str(value).strip():
        return ()
    codes = [code.strip() for code in str(value).split(";") if code.strip()]
    seen: set[str] = set()
    deduped: list[str] = []
    for code in codes:
        if code not in seen:
            seen.add(code)
            deduped.append(code)
    return tuple(deduped)


def _parse_quotation(value: object) -> int | None:
    if value is None or not str(value).strip():
        return None
    if isinstance(value, bool):
        raise ValidationError("Player quotation must be a non-negative integer")
    try:
        quotation = int(value)
    except (TypeError, ValueError) as exc:
        raise ValidationError("Player quotation must be a non-negative integer") from exc
    if (isinstance(value, float) and value != quotation) or quotation < 0:
        raise ValidationError("Player quotation must be a non-negative integer")
    return quotation
